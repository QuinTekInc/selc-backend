from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet

from . import report_commons
from .bulk_report import BulkExcelReport

from selc_core.models import Department, Lecturer, ClassCourse


class DepartmentalExcelReport:

    def __init__(self, department: Department) -> None:
        self.department = department

        #get the lectures in the department
        self.lecturers = Lecturer.objects.filter(department=department)

        #get the class_courses offered by lectures in the department
        self.class_courses = ClassCourse.getCurrentClassCourses().filter(lecturer__in=self.lecturers)

        #create a work book here.
        self.work_book = report_commons.create_workbook()

        self.overview_sheet()

        #create the common bulk report.
        self.bulk_report = BulkExcelReport(self.work_book, self.class_courses)

        pass

    def overview_sheet(self):

        from openpyxl.styles import Font

        ws = self.work_book.create_sheet(title='Overview')

        col_widths = {1: 35, 2: 50}

        report_commons.set_column_widths(ws, col_widths)

        report_commons.init_sheet_title(ws, title='Course Evaluation Overview (Departmental Level)', span_column=2)

        data = [
            ('Department Name', self.department.department_name),
            ('Number of Lecturers', len(self.lecturers)),
            ('Academic Year', 'Academic Year here'),
            ('Semester', 'Semester here')
        ]

        row = 3

        for d in data:
            for col, d_item in enumerate(d, start=1):
                field_cell = ws.cell(row=row, column=col, value= d_item)

                if col == 1:
                    field_cell.font = Font(bold=True,)
                    field_cell.alignment = Alignment(horizontal='left', vertical='center')
                    pass

                pass

            row += 1
            pass

        pass


    def save(self):
        file_name = f'{self.department.department_name}_dept_report'
        file_type = '.xlsx'

        report_commons.saveWorkbook(self.work_book, file_name, file_type)

    pass

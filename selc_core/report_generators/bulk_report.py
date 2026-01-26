from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment

from . import report_commons

from selc_core.models import ClassCourse, QuestionCategory


"""
This class is actually a helper class (Not necessarily a stand alone class)
Used to creating excel reports that requires multiple information
Eg: Excel report for evaluation at the administrator level or the department level
"""
class BulkExcelReport:

    def __init__(self, work_book: Workbook, class_courses: list):

        self.work_book: Workbook = work_book

        self.class_courses: list = class_courses

        self.questionnaire_answer_sheet()
        self.categories_score_sheet()
        self.lecturer_rating_sheet()
        self.response_rate_sheet()
        self.suggestion_sentiments_summary_sheet()
        pass


    def questionnaire_answer_sheet(self):

        ws = self.work_book.create_sheet(title='Questionnaire Answer Summary')

        # set the column widths
        column_widths = {1: 20, 2: 50, 3: 10, 4: 10, 5: 15, 6: 10, 7: 15, 8: 15, 9: 15, 10: 20}
        report_commons.set_column_widths(ws, column_widths)

        report_commons.init_sheet_title(ws, title='Quality Assurance and Academic Planning Directorate', span_column=10)
        report_commons.init_sheet_title(ws, title='University of Energy And natural Resource', row=2, span_column=10)

        row: int = 4

        for class_course in self.class_courses:
            new_row = self.build_questionnaire_evaluation_cell(ws, class_course, row)

            ws.merge_cells(start_row=new_row + 1, start_column=1, end_row=new_row + 4, end_column=10)

            row = new_row + 5
            pass

        pass


    def categories_score_sheet(self):

        ws = self.work_book.create_sheet(title='Categories Scores')

        headers = [
            'Lecturer',
            'Department',
            'Level',
            'Course Code',
            'Course Title',
            'Thematic Areas (Categories)'
        ]

        category_names = [category.cat_name for category in QuestionCategory.objects.all()]

        subheaders = []

        for cat_name in category_names:
            subheaders.extend([cat_name, 'Remark'])

        span_col = len(headers) + len(subheaders) - 1

        report_commons.init_sheet_title(ws,
                                        title='Thematic Areas of Evaluation(Category) and their ratings for each lecturer',
                                        span_column=span_col)

        report_commons.init_header_cells(ws, headers)

        report_commons.init_header_cells(ws, subheaders, row=3, start_col=len(headers))
        report_commons.set_row_height(ws, 2, 0.6, in_inches=True)


        # todo: merge the lecturer, department, course code, title header cells vertically
        for header_col, _ in enumerate(headers[0: len(headers) - 1], start=1):
            ws.merge_cells(start_row=2, start_column=header_col, end_row=3, end_column=header_col)
            pass

        # todo: merge the header cell horizontally to be at par with the sub header cells
        ws.merge_cells(start_row=2, start_column=len(headers), end_row=2, end_column=len(headers) + len(subheaders) - 1)

        col_widths = {1: 50, 2: 45, 3: 15,  4: 25, 5: 45, }

        w_skip = False
        for col in range(6, span_col + 1):

            if not w_skip:
                col_widths[col] = 40
            else:
                col_widths[col] = 20

            w_skip = not w_skip

        report_commons.set_column_widths(ws, col_widths)

        # todo: load the actual data

        row = 4  # we start from because while last row index while creating the headers was 3

        # populate the data header.
        for class_course in self.class_courses:
            cc_map = class_course.toMap()

            lecturer_name = class_course.lecturer.getFullName()
            department = class_course.lecturer.department.department_name
            level = class_course.level
            course_code = class_course.course.course_code
            course_title = class_course.course.title

            ws.cell(row=row, column=1, value=lecturer_name)

            ws.cell(row=row, column=2, value=department)

            ws.cell(row=row, column=3, value=level)

            ws.cell(row=row, column=4, value=course_code)

            ws.cell(row=row, column=5, value=course_title)

            categories_list = class_course.getEvalDetails()

            col = 6

            for category_item in categories_list:
                cat_mean_score = category_item['mean_score']
                cat_remark = category_item['remark']

                ws.cell(row=row, column=col, value=cat_mean_score)
                ws.cell(row=row, column=col + 1, value=cat_remark)

                col += 2
                pass

            row += 1
            pass

        pass

    def lecturer_rating_sheet(self):
        ws = self.work_book.create_sheet(title='Lecturer Rating Summary Report')

        headers = [
            'Lecturer',
            'Department',
            'Level',
            'Course Code',
            'Course Title',
            'Ratings'
        ]

        ratings = [i for i in range(5, 0, -1)]

        subheaders = []

        for rating in ratings:
            subheaders.extend([rating, 'Percentage(%)'])

        span_col = len(headers) + len(subheaders) - 1

        # initialize the worksheet title
        report_commons.init_sheet_title(ws, title='Ratings Summary for every lecturer', span_column=span_col)

        # initialize headers and subheaders
        report_commons.init_header_cells(ws, headers)

        report_commons.init_header_cells(ws, subheaders, row=3, start_col=len(headers))
        report_commons.set_row_height(ws, 2, 0.3, in_inches=True)

        # span the last main header col to be at par with the subheaders.
        ws.merge_cells(start_row=2, start_column=len(headers), end_row=2, end_column=span_col)

        for header_col, _ in enumerate(headers[0: len(headers) - 1], start=1):
            ws.merge_cells(start_row=2, start_column=header_col, end_row=3, end_column=header_col)

        col_widths = {1: 50, 2: 45, 3: 20,  4: 25, 5: 45}

        for col in range(6, span_col + 1):
            col_widths[col] = 25

        report_commons.set_column_widths(ws, col_widths)

        # load actual data
        row = 4

        for class_course in self.class_courses:
            lecturer_name = class_course.lecturer.getFullName()
            department = class_course.lecturer.department.department_name
            level = class_course.level
            course_code = class_course.course.course_code
            course_title = class_course.course.title

            rating_summary = class_course.getLecturerRatingDetails()

            ws.cell(row=row, column=1, value=lecturer_name)
            ws.cell(row=row, column=2, value=department)
            ws.cell(row=row, column=3, value=level)
            ws.cell(row=row, column=4, value=course_code)
            ws.cell(row=row, column=5, value=course_title)

            col = 6

            for rating_item in rating_summary:
                rating_count = rating_item['rating_count']
                rating_percentage = rating_item['percentage']

                ws.cell(row=row, column=col, value=rating_count)
                ws.cell(row=row, column=col + 1, value=rating_percentage)

                col += 2
                pass

            row += 1
            pass

        pass

    def response_rate_sheet(self):

        ws = self.work_book.create_sheet(title='Response Rate Summary')

        headers = [
            'Lecturer',
            'Department',
            'Level',
            'Course Code',
            'Course Title',
            'Number of Students',
            'Number of Respondents',
            'Response Rate'
        ]

        report_commons.init_sheet_title(ws, title='Response rate to course evaluations for various lecturers',
                                        span_column=len(headers))

        report_commons.init_header_cells(ws, headers)
        report_commons.set_row_height(ws, 2, 0.4, in_inches=True)


        col_widths = {1: 50, 2: 45, 3: 30, 4: 45, 5: 35, 6: 35, 7: 38}
        report_commons.set_column_widths(ws, col_widths)

        row = 3

        for class_course in self.class_courses:
            cc_map = class_course.toMap()
            lecturer_name = cc_map['lecturer']['name']
            department = cc_map['lecturer']['department']
            level = cc_map['level']
            course_code = cc_map['course']['course_code']
            course_title = cc_map['course']['course_title']
            number_of_students = cc_map['number_of_registered_students']
            number_of_evaluated = cc_map['number_of_evaluated_students']
            response_rate = cc_map['response_rate']

            ws.cell(row=row, column=1, value=lecturer_name)
            ws.cell(row=row, column=2, value=department)
            ws.cell(row=row, column=3, value=level)
            ws.cell(row=row, column=5, value=course_code)
            ws.cell(row=row, column=6, value=course_title)

            ws.cell(row=row, column=7, value=number_of_students)
            ws.cell(row=row, column=8, value=number_of_evaluated)
            ws.cell(row=row, column=9, value=response_rate)

            row += 1

            pass

        pass

    def suggestion_sentiments_summary_sheet(self):

        ws = self.work_book.create_sheet(title='Suggestion Sentiments')

        headers = [
            'Lecturer',
            'Department',
            'Level',
            'Course Code',
            'Course Title',
            'Suggestion Sentiments'
        ]

        sentiments = ['Negative', 'Neutral', 'Positive']

        subheaders = []

        for sentiment in sentiments:
            subheaders.extend([sentiment, 'Percentage(%)'])
            pass

        span_col = len(headers) + len(subheaders) - 1

        report_commons.init_sheet_title(ws, title='Suggestion Summary Report for every lecturer', span_column=span_col)

        report_commons.init_header_cells(ws, headers)

        report_commons.init_header_cells(ws, subheaders, row=3, start_col=len(headers))
        report_commons.set_row_height(ws, 3, 0.4, in_inches=True)

        for col, _ in enumerate(headers[0: len(headers) - 1], start=1):
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            pass

        col_widths = {1: 50, 2: 45, 3: 15,  4: 25, 5: 45}

        for col in range(5, span_col + 1):
            col_widths[col] = 25

        report_commons.set_column_widths(ws, col_widths)

        # load the actual data
        row = 4

        ws.merge_cells(start_row=2, start_column=len(headers), end_row=2, end_column=span_col)
        # todo: populate the sentiment data
        for class_course in self.class_courses:
            lecturer_name = class_course.lecturer.getFullName()
            department = class_course.lecturer.department.department_name
            course_code = class_course.course.course_code
            course_title = class_course.course.title
            sentiment_summary = class_course.getEvalSuggestions(include_suggestions=False)['sentiment_summary']

            ws.cell(row=row, column=1, value=lecturer_name)
            ws.cell(row=row, column=2, value=department)
            ws.cell(row=row, column=3, value=course_code)
            ws.cell(row=row, column=4, value=course_title)

            col = 6

            for sentiment_item in sentiment_summary:
                sentiment_count = sentiment_item['sentiment_count']
                sentiment_percent = sentiment_item['sentiment_percent']

                ws.cell(row=row, column=col, value=sentiment_count)
                ws.cell(row=row, column=col + 1, value=sentiment_percent)

                col += 2
                pass

            row += 1
            pass

        pass

    def build_questionnaire_evaluation_cell(self, sheet: Worksheet, class_course: ClassCourse, q_start_row) -> int:

        current_row = q_start_row

        cc_map = class_course.toMap()

        # create the detail cells

        """
        Name of Lecturer: name  ------- Semester: semester
        Course title: title ------ Course Code: code ---- credit_hours: credit_hours
        Number of Students in Class: number of students ----------  Number of respondents: credit ---- Response rate: response rate
        Department roll: number on roll ----- Lecturer rating for course: l_rating
        course_score: average score ---- percentage score: percentage ----- remark: generated remark

        """

        # name detail cells
        self.build_field_cell(sheet, row=current_row, column=1, value='Name of Lecturer')

        sheet.cell(row=current_row, column=2, value=class_course.lecturer.getFullName())  # spans 3 columns BCD
        # perform necessary row merging
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        # semester detail cells
        self.build_field_cell(sheet, row=current_row, column=5, value='Semester')

        sheet.cell(row=current_row, column=6, value=class_course.semester)

        #academic year detail cells
        self.build_field_cell(sheet, row=current_row, column=7, value='Year')

        sheet.cell(row=current_row, column=8, value=class_course.year)

        # Level detail cells
        self.build_field_cell(sheet, row=current_row, column=9, value='Level')

        sheet.cell(row=current_row, column=10, value=class_course.level)

        #sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=10)

        # ==================next row=========================================

        current_row += 1

        # course title cells
        self.build_field_cell(sheet, row=current_row, column=1, value='Course Title')

        sheet.cell(row=current_row, column=2, value=class_course.course.title)  # spans 3 columns BCD
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        # course code cells
        self.build_field_cell(sheet, row=current_row, column=5, value='Course Code')

        sheet.cell(row=current_row, column=6, value=class_course.course.course_code)  # span 3 columns FGH
        sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=8)

        # credit hour cells
        self.build_field_cell(sheet, row=current_row, column=9, value='Credit Hours')

        sheet.cell(row=current_row, column=10, value=class_course.credit_hours)

        # =================next row============================================

        current_row += 1

        # number on roll cells
        self.build_field_cell(sheet, row=current_row, column=1, value='Number of Students')

        sheet.cell(row=current_row, column=2, value=cc_map['number_of_registered_students'])  # spans 3 columns BCD
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        # respondent cells
        self.build_field_cell(sheet, row=current_row, column=5, value='Number of respondents')

        sheet.cell(row=current_row, column=6, value=cc_map['number_of_evaluated_students'])
        sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=8)

        # response rate cells
        self.build_field_cell(sheet, row=current_row, column=9, value='Response Rate')

        sheet.cell(row=current_row, column=10, value=cc_map['response_rate'])

        # ==================next row=========================================

        current_row += 1

        # department cells
        self.build_field_cell(sheet, row=current_row, column=1, value='Department')

        sheet.cell(row=current_row, column=2,
                   value=class_course.lecturer.department.department_name)  # spans 3 columns BCD
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        # Lecturer rating for course cells
        self.build_field_cell(sheet, row=current_row, column=5, value='Lecturer rating for this course')

        sheet.cell(row=current_row, column=6, value=cc_map['lecturer_course_rating'])  # spans 5 columns FGHIJ
        sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=10)

        # ==================next row=========================================

        current_row += 1

        # Course score cell
        self.build_field_cell(sheet, row=current_row, column=1, value='Mean Score')

        sheet.cell(row=current_row, column=2, value=cc_map['grand_mean_score'])  # spans 3 columns BCD
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

        # percentage course score cells
        self.build_field_cell(sheet, row=current_row, column=5, value='Percentage Score')

        sheet.cell(row=current_row, column=6, value=cc_map['grand_percentage'])  # spans 3 columns FGH
        sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=8)

        # remarks
        self.build_field_cell(sheet, row=current_row, column=9, value='Remark')

        sheet.cell(row=current_row, column=10, value=cc_map['remark'])


        #todo: add a new row to show the campus, week/mainstream session, number of programs (specifically for general courses)
        # don't forget to increase the current_row variable first


        #set the row heights for the class_course detail fields
        for r in range(q_start_row, current_row + 1):
            report_commons.set_row_height(sheet, r, 0.4, in_inches=True)
            pass


        # ======================next row=======================================================
        current_row += 1

        # todo: populating the headers [header title, span]
        headers = [
            'Core Area (Categories)',
            'Question',
            'Q Mean Score',
            'Q Percentage Score',
            'C Mean Score',
            'C Percentage Score',
            'C Remark',
            'Grand Mean Score',
            'Percentage Score',
            'Remark'
        ]

        report_commons.init_header_cells(sheet, headers, row=current_row)
        report_commons.set_row_height(sheet, current_row, 0.6, in_inches=True)

        # for populating the actual categories and questions data
        current_row += 1

        row = current_row

        # get the evaluation
        evaluation_summary = class_course.getEvalDetails()

        # loop for categories
        for eval_item in evaluation_summary:

            # get the category
            category = eval_item['category']
            cat_mean_score = eval_item['mean_score']
            cat_percentage_score = eval_item['percentage_score']
            cat_remark = eval_item['remark']

            questions: list = eval_item['questions']

            sheet.cell(row=row, column=1, value=category)

            start_row = row
            end_row = start_row + len(questions) - 1

            # inner loop for populating the questions
            for question_item in questions:
                question = question_item['question']
                q_mean_score = question_item['mean_score']
                q_percentage_score = question_item['percentage_score']

                report_commons.create_cell(sheet, row=row, column=2, value=question)
                sheet.cell(row=row, column=3, value=q_mean_score)
                sheet.cell(row=row, column=4, value=q_percentage_score)

                if (questions.index(question_item) == len(questions) - 1 and
                        evaluation_summary.index(eval_item) == len(evaluation_summary) - 1):
                    break

                row += 1
                pass

            sheet.cell(row=start_row, column=5, value=cat_mean_score)
            sheet.cell(row=start_row, column=6, value=cat_percentage_score)
            sheet.cell(row=start_row, column=7, value=cat_remark)


            # todo: merge necessary cells
            # category cell
            report_commons.merge_cells(sheet, start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            # mean score cell
            report_commons.merge_cells(sheet, start_row=start_row, start_column=5, end_row=end_row, end_column=5)
            # percentage score cell
            report_commons.merge_cells(sheet, start_row=start_row, start_column=6, end_row=end_row, end_column=6)
            # category remark cell
            report_commons.merge_cells(sheet, start_row=start_row, start_column=7, end_row=end_row, end_column=7)

            pass

        # grand mean score
        sheet.cell(row=current_row, column=8, value=cc_map['grand_mean_score'])
        report_commons.merge_cells(sheet, center=True, start_row=current_row, start_column=8, end_row=row, end_column=8)

        # grand percentage score
        sheet.cell(row=current_row, column=9, value=cc_map['grand_percentage'])
        report_commons.merge_cells(sheet, center=True, start_row=current_row, start_column=9, end_row=row, end_column=9)

        # final remark
        sheet.cell(row=current_row, column=10, value=cc_map['remark'])
        report_commons.merge_cells(sheet, center=True, start_row=current_row, start_column=10, end_row=row, end_column=10)

        return row  # the last row  after the function operation

    def build_field_cell(self, sheet: Worksheet, row: int, column: int, value):
        field_cell = sheet.cell(row=row, column=column, value=value)
        field_cell.font = Font(bold=True, size=11)
        field_cell.alignment = Alignment(vertical='center', horizontal='left', wrapText=True, )

        return field_cell

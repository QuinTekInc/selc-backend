
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from io import BytesIO

from django.core.files.base import ContentFile
from selc_core.models import ReportFile

from django.db.transaction import  atomic as atomic_transaction




def create_workbook():
    work_book = Workbook()
    #get the default active sheet.
    active_sheet = work_book.active
    work_book.remove(active_sheet)

    return work_book



#for creating title text normally at the first row of every work sheet.
def init_sheet_title(sheet: Worksheet, title='', row=1, column=1, span_column=None):
    header_cell = sheet.cell(row=row, column=column, value=title)
    header_cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    header_cell.font = Font(bold=True, size=14)

    if span_column is not None:
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=span_column)

    return header_cell




#function that populates table headers in a given worksheet.
def init_header_cells(sheet: Worksheet, headers: list, row=2, start_col=1):

    if not headers:
        return

    for col, header in enumerate(headers, start=start_col):
        header_cell = sheet.cell(row=row, column=col, value=header)
        header_cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        header_cell.font = Font(bold=True, size=12)
        pass

    pass




def set_column_widths(sheet: Worksheet, widths: dict):
    """
    widths = {column_index: width_value}
    """
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = width
        pass
    pass


def set_row_height(sheet: Worksheet, row:int, height: float, in_inches=False):

    if in_inches:
        #function to convert the height from inches to pixels
        height = inch_to_point(height)

    sheet.row_dimensions[row].height = height
    pass


def inch_to_point(num):
    return num * 72 # convert the number from inches into points (1 inch = 72 points) or (1 point = 1/72 inch)



def create_cell(work_sheet: Worksheet, row: int, column: int, value=None):

    cell = work_sheet.cell(row=row, column=column, value=value)


    #align the content or value of cells to right for numbers and to left for strings

    if isinstance(value, int) or isinstance(value, float):
        cell.alignment = Alignment(vertical='top', horizontal='right', wrap_text=True)
        pass
    elif isinstance(value, str):
        cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
        pass

    return cell



def merge_cells(work_sheet: Worksheet, start_row: int, start_column: int, end_row: int, end_column: int, alignment=None, center=False):

    cell = work_sheet.cell(row=start_row, column=start_column)

    if alignment is not None:
        cell.alignment = alignment

    #this is the default alignment
    if alignment is None:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    #perform the actual merging of the cells
    work_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    pass




def saveWorkbook(work_book: Workbook, file_name: str, file_type: str) -> ReportFile:

    file_stream = BytesIO()
    work_book.save(file_stream)
    file_stream.seek(0)

    content_file = ContentFile(file_stream.read())

    with atomic_transaction():
        report_file, created = ReportFile.objects.update_or_create(file_name=file_name, file_type=file_type)

        if not created:
            report_file.file.delete(save=False)

        report_file.file.save(f'{file_name}{file_type}', content_file)

    content_file.close()
    file_stream.close()

    return report_file




def savePdf(file_stream: BytesIO, file_name: str, file_type: str) -> ReportFile:


    content_file = ContentFile(file_stream.read())

    with atomic_transaction():
        report_file, created = ReportFile.objects.update_or_create(file_name=file_name, file_type=file_type)

        if not created:
            report_file.file.delete(save=False)

        report_file.file.save(f'{file_name}{file_type}', content_file)

    content_file.close()
    file_stream.close()
        
    return report_file



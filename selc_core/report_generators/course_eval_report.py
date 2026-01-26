from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook

from selc_core.core_utils import getScoreRemark
from selc_core.models import ClassCourse

from . import report_commons


from io import BytesIO
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    Spacer,
    PageBreak,
    Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_LEFT
from reportlab.lib import colors

from pathlib import Path



class CourseEvalExcelReport:

    def __init__(self, class_course: ClassCourse):

        self.class_course = class_course

        self.lecturer = class_course.lecturer

        self.work_book = Workbook()

        #remove the default work sheet.
        self.work_book.remove(self.work_book.active)

        #load the category info
        self.eval_summary = self.class_course.getEvalDetails()

        self.overview_sheet()
        self.questionnaire_answer_summary_sheet()
        self.category_scores_sheet()
        self.lecturer_rating_summary_sheet()
        self.suggestion_sentiments_summary_sheet()

        pass

    def overview_sheet(self):
        ws = self.work_book.create_sheet(title='Overview')

        number_of_students, evaluated_students = self.class_course.getNumberOfRegisteredStudents()
        response_rate = (evaluated_students / number_of_students) * 100

        course_mean_score, percentage_score = self.class_course.computeGrandMeanScore()

        # lecturer name
        # department
        # course code
        # course title
        # credit hours
        # semester
        # academic year
        # number of registered students
        # evaluated students
        # response rate
        # course score
        # percentage score
        # average sentiment
        # lecturer rating for this course.

        programs = self.class_course.getListOfProgramsInClass()

        overview_list = [
            ('Lecturer Name', self.class_course.lecturer.getFullName()),
            ('Department', self.class_course.lecturer.department.department_name),
            ('Course code', self.class_course.course.course_code),
            ('Course title', self.class_course.course.title),
            ('Credit Hours', self.class_course.credit_hours),
            ('Level', self.class_course.level),
            ('Semester', self.class_course.semester),
            ('Academic year', self.class_course.year),
            ('Number of Students', number_of_students),
            ('Evaluated Students', evaluated_students),
            ('Response Rate', response_rate),
            ('Mean Score', course_mean_score),
            ('Percentage Score', percentage_score),
            ('Remark', getScoreRemark(course_mean_score)),
            ('Lecturer Rating for this course', self.class_course.computeLecturerRatingForCourse()),
            ('Number of Programs', len(programs))
        ]

        report_commons.init_sheet_title(ws, title='Course Evaluation Overview', span_column=2)

        report_commons.set_column_widths(ws, {1: 30, 2: 50})

        current_row_index = 2

        for row, overview_field in enumerate(overview_list, start=2):
            field_cell = ws.cell(row=row, column=1, value=overview_field[0])
            field_cell.alignment = Alignment(horizontal='center', vertical='center')
            field_cell.font = Font(bold=True)

            #data cell
            ws.cell(row=row, column=2, value=overview_field[1])

            current_row_index += 1
            pass

        #add the included programs
        ws.cell(row=current_row_index, column=1, value='Programs')

        for row, program in enumerate(programs, start=current_row_index):
            ws.cell(row=row, column=1, value=program)
            pass

        pass

    def questionnaire_answer_summary_sheet(self):

        ws = self.work_book.create_sheet(title='Questionnaire Answers')

        headers = ['Question', 'Answer Type', 'Answer Option', 'Count', 'Average Score', 'Percentage Score', 'Remark']

        report_commons.init_sheet_title(ws, title='Questionnaire Answer Summary', span_column=len(headers))

        #create the headers.
        report_commons.init_header_cells(ws, headers=headers)
        report_commons.set_row_height(ws, 2, 0.4, in_inches=True)

        report_commons.set_column_widths(ws, {1: 50, 2: 10, 3: 10, 4: 12, 5: 15, 6: 15, 7: 10})

        # extract the questionnaire answer summaries
        questions_answer_summary: list = [question_dict for summary in self.eval_summary for question_dict in
                                          summary.get('questions', [])]
        '''
            {
              'question': 'Coverage of course content by lecturer',
              'answer_type': 'performance',
              'answer_summary': {
                'Poor': 0,
                'Average': 0,
                'Good': 0,
                'Very Good': 1,
                'Excellent': 0,
              },
              'percentage_score': 80.0,
              'average_score': 4.0,
              'remark': 'Very Good'
            },
            {
              'question': 'Communication of objectives of the course',
              'answer_type': 'performance',
              'answer_summary': {
                'Poor': 0,
                'Average': 0,
                'Good': 0,
                'Very Good': 1,
                'Excellent': 0,
              },
              'percentage_score': 80.0,
              'average_score': 4.0,
              'remark': 'Very Good'
            }
        '''

        row = 3

        for eval_summary_item in questions_answer_summary:

            question: str = eval_summary_item['question']
            answer_type: str = eval_summary_item['answer_type']
            answer_summary: dict = eval_summary_item['answer_summary']
            average_score: float = eval_summary_item['mean_score']
            percentage_score: float = eval_summary_item['percentage_score']
            remark: float = eval_summary_item['remark']

            #populate the question cell.
            report_commons.create_cell(ws, row=row, column=1, value=question)

            ws.cell(row=row, column=2, value=answer_type)

            start_row = row
            end_row = row + len(answer_summary) - 1

            for answer_entry in answer_summary.items():
                option, count = answer_entry

                ws.cell(row=row, column=3, value=option)
                ws.cell(row=row, column=4, value=count)

                row += 1
                pass

            ws.cell(row=start_row, column=5, value=average_score)

            ws.cell(row=start_row, column=6, value=percentage_score)

            ws.cell(row=start_row, column=7, value=remark)

            #todo: span the various cells apart to at the par with the cells of its corresponding answer_option and count cells.

            #spanning the question cell
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

            # spanning the answer_type cell
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

            # spanning the average score cell
            ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)

            # spanning the percentage score cell
            ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)

            # spanning the remark cell.
            ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)
            pass

        pass

    def category_scores_sheet(self):

        ws = self.work_book.create_sheet(title='Category Summary')

        headers = ['Core Area (Category)', 'Questions', 'Average Score', 'Percentage Score', 'Remark']

        report_commons.init_sheet_title(ws, title='Thematic Areas of Evaluation (Categories)', span_column=len(headers))

        report_commons.init_header_cells(ws, headers)
        report_commons.set_row_height(ws, 2, 0.4, in_inches=True)

        report_commons.set_column_widths(ws, {1: 30, 2: 50, 3: 15, 4: 15, 5: 10})

        category_summary: list = self.class_course.getEvalQuestionCategoryRemarks(include_questions=True)

        row = 3

        #todo: populate the categories scores here.
        for summary_item in category_summary:
            category: str = summary_item['category']
            questions: list = [question['question'] for question in summary_item['questions']]
            average_score: float = summary_item['average_score']
            percentage_score: float = summary_item['percentage_score']
            remark: str = summary_item['remark']

            ws.cell(row=row, column=1, value=category)

            start_row = row
            end_row = row + len(questions) - 1

            for question in questions:
                report_commons.create_cell(ws, row=row, column=2, value=question)
                row += 1
                pass

            ws.cell(row=start_row, column=3, value=average_score)
            ws.cell(row=start_row, column=4, value=percentage_score)
            ws.cell(row=start_row, column=5, value=remark)

            #todo: span the cells vertically
            #category cell
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            #average score cell
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
            #percentage score cell
            ws.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
            #remark
            ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
            pass

        pass

    def lecturer_rating_summary_sheet(self):

        ws = self.work_book.create_sheet(title='Lecturer Rating')

        headers = ['Rating', 'Count', 'Percentage']

        report_commons.init_sheet_title(ws, title='Lecturer Rating Summary', span_column=len(headers))

        report_commons.init_header_cells(ws, headers)
        report_commons.set_row_height(ws, 2, 0.3, in_inches=True)

        report_commons.set_column_widths(ws, {1: 10, 2: 15, 3: 17})

        lecturer_rating_summary = self.class_course.getLecturerRatingDetails()

        row = 3

        for rating_summary in lecturer_rating_summary:
            rating = rating_summary['rating']
            count = rating_summary['rating_count']
            percentage = rating_summary['percentage']

            ws.cell(row=row, column=1, value=rating)

            ws.cell(row=row, column=2, value=count)

            ws.cell(row=row, column=3, value=percentage)

            row += 1
            pass

        pass

    def suggestion_sentiments_summary_sheet(self):

        ws = self.work_book.create_sheet(title='Sentiment Summary')

        headers = ['Sentiment', 'Count', 'Percentage']

        report_commons.init_sheet_title(ws, title='Suggestions Sentiment Summary', span_column=len(headers))

        report_commons.init_header_cells(ws, headers)
        report_commons.set_row_height(ws, 2, 0.3, in_inches=True)

        report_commons.set_column_widths(ws, {1: 20, 2: 15, 3: 17})

        sentiment_summary: list = self.class_course.getEvalSuggestions(include_suggestions=False)['sentiment_summary']

        row = 3

        for sentiment_summary_item in sentiment_summary:
            sentiment: str = sentiment_summary_item['sentiment']
            count: int = sentiment_summary_item['sentiment_count']
            percentage: float = sentiment_summary_item['sentiment_percent']

            ws.cell(row=row, column=1, value=sentiment)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=percentage)

            row += 1
            pass

        pass

    #saves generated report
    #returns a ReportFile database object
    def save(self):
        file_name = self.class_course.getSavableReportFileName()
        file_type = '.xlsx'

        return report_commons.saveWorkbook(self.work_book, file_name, file_type)

    pass







class CourseEvalPdfReport:
    HEADER_BG = colors.whitesmoke
    GRID_COLOR = colors.lightgrey

    def __init__(self, class_course: ClassCourse):

        self.class_course = class_course
        self.file_buffer = BytesIO()
        self.styles = getSampleStyleSheet()
        self.story = []

        # 🔹 Table text style (WRAPPING ENABLED)
        self.table_text = ParagraphStyle(
            name="TableText",
            fontName="Helvetica",
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
        )

        self.table_header = ParagraphStyle(
            name="TableHeader",
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            alignment=TA_LEFT,
        )

        self.document = SimpleDocTemplate(
            self.file_buffer,
            pagesize=A4,
            rightMargin=16,
            leftMargin=16,
            topMargin=16,
            bottomMargin=16,
        )

        self.init_pages()

    # ------------------------------------------------------------------
    # Page builder
    # ------------------------------------------------------------------
    def init_pages(self):
        self.report_header()
        self.course_information()

        eval_details = self.class_course.getEvalDetails()

        self.section("Questionnaire Summary")
        self.paragraph(
            "Summary information on how students answered the evaluation questionnaire."
        )
        self.questionnaire_table(eval_details)

        self.section("Summary Report")
        self.paragraph(
            "The table below shows the categorical (core area) summary of the questionnaire."
        )
        self.category_table(eval_details)

        self.section("Remarks Reference")
        self.paragraph(
            "This table shows the range of mean/average scores and their corresponding remarks."
        )
        self.scoring_table()

        self.story.append(PageBreak())

        self.section("Lecturer Rating Summary")
        self.paragraph(
            "This table shows the rating summary of the lecturer for this course."
        )
        self.lecturer_rating_table()

        self.section("Suggestion Sentiment Summary")
        self.paragraph(
            "This shows the summary of sentiments of students regarding this course and lecturer."
        )
        self.sentiment_table()

    # ------------------------------------------------------------------
    # Header & helpers
    # ------------------------------------------------------------------
    def report_header(self):

        logo_path = f'{Path(__file__).parent}/UENR-Logo.png'

        #todo: add the university logo to the left side of the header
        logo = Image(logo_path, width=60, height=60)

        text = [
            Paragraph("<b>University of Energy and Natural Resources</b>", self.styles["Title"]),
            Paragraph("Quality Assurance and Academic Planning Directorate", self.styles["Heading2"]),
            Paragraph("Students Evaluation of Lecturers and Courses (SELC)", self.styles["Heading3"]),
        ]

        #todo: replace the empty string with the logo image when available
        table = Table([[logo, text]], colWidths=[70, 450])
        table.setStyle([("VALIGN", (0, 0), (-1, -1), "TOP")])
        self.story.extend([table, Spacer(1, 16)])

    def section(self, title):
        table = Table([[Paragraph(title, self.table_header)]], colWidths=[520])
        table.setStyle([
            ("BACKGROUND", (0, 0), (-1, -1), self.HEADER_BG),
            ("BOX", (0, 0), (-1, -1), 0.5, self.GRID_COLOR),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])
        self.story.extend([table, Spacer(1, 8)])

    def paragraph(self, text):
        self.story.extend([Paragraph(text, self.styles["Normal"]), Spacer(1, 8)])

    # ------------------------------------------------------------------
    # Tables (WRAPPED)
    # ------------------------------------------------------------------
    def course_information(self):
        self.section("Course Information and Evaluation Overview")

        cc_map = self.class_course.toMap()

        registered_students = cc_map['number_of_registered_students']
        evaluated_students = cc_map['number_evaluated_students']

        response_rate = (evaluated_students / registered_students) * 100 if registered_students > 0 else 0

        rows = [
            ('FIELD', 'VALUE'),  #this is the table header
            ("Lecturer", self.class_course.lecturer.getFullName()),
            ("Department", self.class_course.lecturer.department.department_name),
            ("Email", self.class_course.lecturer.user.email),
            ("Course Code", self.class_course.course.course_code),
            ("Course Title", self.class_course.course.title),
            ('Credit Hours:', self.class_course.credit_hours),
            ('Level', self.class_course.level),
            ('Programs', len(self.class_course.getListOfProgramsInClas())),
            ('Semester', cc_map['semester']),
            ("Year", cc_map['year']),
            ("Registered Students", registered_students),
            ("Evaluated Students", evaluated_students),
            ("Response Rate", f"{response_rate:.2f}%"),
            ("Course Mean Score", cc_map['grand_mean_score']),
            ("Course Percentage Score", cc_map['grand_percentage']),
            ('Remark', cc_map['remark'])
        ]

        data = [
            [
                Paragraph(f"{k}:", self.table_header),
                Paragraph(str(v), self.table_text),
            ]
            for k, v in rows
        ]

        self.standard_table(data, [150, 350])

    def questionnaire_table(self, eval_details):
        data = [[
            Paragraph("Question", self.table_header),
            Paragraph("Answer Summary", self.table_header),
            Paragraph("Mean", self.table_header),
            Paragraph("Percentage (%)", self.table_header),
            Paragraph("Remark", self.table_header),
        ]]

        for detail in eval_details:
            for q in detail["questions"]:
                answers = "<br/>".join(
                    f"{k}: {v}" for k, v in q["answer_summary"].items()
                )

                data.append([
                    Paragraph(q["question"], self.table_text),
                    Paragraph(answers, self.table_text),
                    Paragraph(str(q["mean_score"]), self.table_text),
                    Paragraph(str(q["percentage_score"]), self.table_text),
                    Paragraph(q["remark"], self.table_text),
                ])

        self.standard_table(data, [160, 140, 60, 80, 80])

    def category_table(self, eval_details):
        data = [[
            Paragraph("Core Area", self.table_header),
            Paragraph("Mean Score", self.table_header),
            Paragraph("Percentage", self.table_header),
            Paragraph("Remark", self.table_header),
        ]]

        for e in eval_details:
            data.append([
                Paragraph(e["category"], self.table_text),
                Paragraph(str(e["mean_score"]), self.table_text),
                Paragraph(str(e["percentage_score"]), self.table_text),
                Paragraph(e["remark"], self.table_text),
            ])

        self.standard_table(data, [200, 100, 100, 120])

    def scoring_table(self):
        rows = [
            ("4.6 - 5.0", "90 - 100", "Excellent"),
            ("4.0 - 4.59", "60 - 89", "Very Good"),
            ("3.0 - 3.99", "40 - 59", "Good"),
            ("2.0 - 2.99", "20 - 39", "Average"),
            ("0.0 - 1.9", "0 - 19", "Poor"),
        ]

        data = [[
            Paragraph("Score Range", self.table_header),
            Paragraph("Percentage Range", self.table_header),
            Paragraph("Remark", self.table_header),
        ]]

        for r in rows:
            data.append([Paragraph(x, self.table_text) for x in r])

        self.standard_table(data, [160, 160, 200])

    def lecturer_rating_table(self):
        data = [[
            Paragraph("Rating", self.table_header),
            Paragraph("Count", self.table_header),
            Paragraph("Percentage", self.table_header),
        ]]

        for r in self.class_course.getLecturerRatingDetails():
            data.append([
                Paragraph(str(r["rating"]), self.table_text),
                Paragraph(str(r["rating_count"]), self.table_text),
                Paragraph(str(r["percentage"]), self.table_text),
            ])

        self.standard_table(data, [160, 160, 160])

    def sentiment_table(self):
        data = [[
            Paragraph("Sentiment", self.table_header),
            Paragraph("Count", self.table_header),
            Paragraph("Percentage", self.table_header),
        ]]

        for s in self.class_course.getEvalSuggestions(False)["sentiment_summary"]:
            data.append([
                Paragraph(s["sentiment"], self.table_text),
                Paragraph(str(s["sentiment_count"]), self.table_text),
                Paragraph(str(s["sentiment_percent"]), self.table_text),
            ])

        self.standard_table(data, [200, 160, 160])

    # ------------------------------------------------------------------
    # Table style
    # ------------------------------------------------------------------
    def standard_table(self, data, col_widths):
        table = Table(data, colWidths=col_widths)
        table.setStyle([
            ("BACKGROUND", (0, 0), (-1, 0), self.HEADER_BG),
            ("GRID", (0, 0), (-1, -1), 0.5, self.GRID_COLOR),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
        self.story.extend([table, Spacer(1, 16)])

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    def save(self):
        self.document.build(self.story)
        self.file_buffer.seek(0)

        return report_commons.savePdf(
            self.file_buffer,
            self.class_course.getSavableReportFileName(),
            ".pdf",
        )


def test():
    from selc_core.models import ClassCourse

    class_course = ClassCourse.objects.first()

    report = CourseEvalPdfReport(class_course)

    report.save()
    pass

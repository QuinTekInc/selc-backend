import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

# Sample data (replace with your real data)
data = [
    {
        'question': 'Coverage of course content by lecturer',
        'answer_type': 'performance',
        'answer_summary': {'Poor': 0, 'Average': 0, 'Good': 0, 'Very Good': 1, 'Excellent': 0},
        'percentage_score': 80.0,
        'average_score': 4.0,
        'remark': 'Very Good'
    },
    {
        'question': 'Communication of objectives of the course',
        'answer_type': 'performance',
        'answer_summary': {'Poor': 0, 'Average': 0, 'Good': 0, 'Very Good': 1, 'Excellent': 0},
        'percentage_score': 80.0,
        'average_score': 4.0,
        'remark': 'Very Good'
    }
]

# Create a workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Survey Results"

# Define headers
headers = ["Question", "Answer Type", "Option", "Count", "Average Score", "Percentage Score", "Remark"]
ws.append(headers)

# Apply bold font to headers
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

# Border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

row_index = 2
for entry in data:
    options = list(entry["answer_summary"].keys())
    counts = list(entry["answer_summary"].values())
    num_options = len(options)

    # Write rows for each option
    for i, (opt, cnt) in enumerate(zip(options, counts)):
        ws.cell(row=row_index + i, column=3, value=opt)
        ws.cell(row=row_index + i, column=4, value=cnt)

    # Merge and fill Question
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + num_options - 1, end_column=1)
    ws.cell(row=row_index, column=1, value=entry["question"])

    # Merge and fill Answer Type
    ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index + num_options - 1, end_column=2)
    ws.cell(row=row_index, column=2, value=entry["answer_type"])

    # Merge and fill Average Score
    ws.merge_cells(start_row=row_index, start_column=5, end_row=row_index + num_options - 1, end_column=5)
    ws.cell(row=row_index, column=5, value=entry["average_score"])

    # Merge and fill Percentage Score
    ws.merge_cells(start_row=row_index, start_column=6, end_row=row_index + num_options - 1, end_column=6)
    ws.cell(row=row_index, column=6, value=entry["percentage_score"])

    # Merge and fill Remark
    ws.merge_cells(start_row=row_index, start_column=7, end_row=row_index + num_options - 1, end_column=7)
    ws.cell(row=row_index, column=7, value=entry["remark"])

    # Add borders and alignment
    for r in range(row_index, row_index + num_options):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = thin_border

    row_index += num_options

# Adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Save workbook
wb.save("survey_results.xlsx")
